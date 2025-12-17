from flask import Flask, render_template, request, jsonify, send_from_directory, redirect, url_for, session
import os
from mieten import fuehre_mietabgleich_durch
import traceback
from datetime import timedelta, datetime
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
RESULTS_FOLDER = os.path.join(BASE_DIR, "results")
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@example.com")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "change-me")

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["RESULTS_FOLDER"] = RESULTS_FOLDER
app.secret_key = os.environ.get("SECRET_KEY", "please-change-me-very-secret")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=60)

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULTS_FOLDER, exist_ok=True)


def process_tours_table(workbook):
    if "touren" not in workbook.sheetnames:
        workbook.create_sheet("touren")
    tours_sheet = workbook["touren"]

    if "ag-grid" not in workbook.sheetnames:
        return
    ag_grid_sheet = workbook["ag-grid"]

    current_date = datetime.utcnow().strftime("%Y-%m-%d")

    tours_sheet.cell(1, 1, "Tour")
    tours_sheet.cell(1, 2, f"TG ({current_date})")

    tour_values = []
    for row in range(2, ag_grid_sheet.max_row + 1):
        tour_value = ag_grid_sheet.cell(row, 1).value
        if tour_value:
            tour_values.append(tour_value)

    unique_tours = sorted(list(set(tour_values)))

    if tours_sheet.max_row >= 2:
        tours_sheet.delete_rows(2, tours_sheet.max_row - 1)

    for i, tour in enumerate(unique_tours, 2):
        tours_sheet.cell(i, 1, tour)

    tour_customer_count = {}
    for row in range(2, ag_grid_sheet.max_row + 1):
        tour = ag_grid_sheet.cell(row, 1).value
        customer = ag_grid_sheet.cell(row, 8).value
        if tour and customer:
            tour_customer_count[tour] = tour_customer_count.get(tour, 0) + 1

    for row_idx in range(2, len(unique_tours) + 2):
        tour = tours_sheet.cell(row_idx, 1).value
        if tour:
            tours_sheet.cell(row_idx, 2, tour_customer_count.get(tour, 0))


def process_excel(file_path, save_directory):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    process_tours_table(workbook)

    blue_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    light_gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    col_ac_idx = 29
    col_ad_idx = 30
    col_ae_idx = 31
    col_af_idx = 32
    col_ag_idx = 33
    col_ah_idx = 34

    for row in range(1, sheet.max_row + 1):
        sheet.cell(row, col_ac_idx).fill = blue_fill
    sheet.cell(1, col_ad_idx).fill = blue_fill

    sheet.cell(1, col_ae_idx).fill = green_fill
    sheet.cell(1, col_af_idx).fill = green_fill

    sheet.cell(1, col_ag_idx).fill = light_gray_fill
    sheet.cell(1, col_ah_idx).fill = light_gray_fill

    values_to_filter = ["D009", "D090", "D091", "D092", "D093", "D094", "D096", "D095", "D208", "D251", "D270", "D271", "D291", "D292", "SCD12", "SCD13"]
    try:
        sheet.auto_filter.add_filter_column(0, values_to_filter)
        new_filter_last_col_letter = get_column_letter(28)
        filter_range = f"A1:{new_filter_last_col_letter}{sheet.max_row}"
        sheet.auto_filter.ref = filter_range
    except Exception:
        pass

    subtotal_source_col_letter = get_column_letter(col_ac_idx)
    current_max_row = sheet.max_row if sheet.max_row >= 2 else 2
    subtotal_range_end_row = max(current_max_row, 2000)
    sheet.cell(1, col_ad_idx).value = f'=SUBTOTAL(9,{subtotal_source_col_letter}2:{subtotal_source_col_letter}{subtotal_range_end_row})'

    sheet.cell(1, col_ae_idx).value = 'Adressen'
    sheet.cell(1, col_af_idx).value = '=SUMPRODUCT(--(FREQUENCY(COLUMN(1:1175),SUBTOTAL(3,INDIRECT("H"&ROW(2:1175)))*MATCH(H2:H1175&"",H2:H1175&"",0))>0))-1'

    sheet.cell(1, col_ag_idx).value = 'Touren'
    sheet.cell(1, col_ah_idx).value = '=SUMPRODUCT(--(FREQUENCY(COLUMN(1:1),SUBTOTAL(3,INDIRECT("A"&ROW(2:1175)))*MATCH(A2:A1175&"",A2:A1175&"",0))>0))-1'

    sheet.column_dimensions['A'].width = 6.5
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 6
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 8
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15

    col_i_idx = 9
    col_j_idx = 10
    sheet.cell(row=1, column=col_i_idx).value = "Ablage"
    sheet.column_dimensions[get_column_letter(col_i_idx)].width = 8
    sheet.cell(row=1, column=col_j_idx).value = "Schlüssel"
    sheet.column_dimensions[get_column_letter(col_j_idx)].width = 8

    col_k_idx = 11
    col_l_idx = 12
    sheet.column_dimensions[get_column_letter(col_k_idx)].width = 15
    sheet.column_dimensions[get_column_letter(col_l_idx)].width = 10

    comment_target_col_idx = col_l_idx
    comment_source_start_col_idx = 13
    comment_source_end_col_idx = 22

    for row_num in range(2, sheet.max_row + 1):
        comment_text = ""
        for col_idx_comment_source in range(comment_source_start_col_idx, comment_source_end_col_idx + 1):
            cell_value = sheet.cell(row_num, col_idx_comment_source).value
            if cell_value is not None:
                column_letter_for_comment = get_column_letter(col_idx_comment_source)
                comment_text += f"{column_letter_for_comment}{row_num}: {cell_value}\n"

        if comment_text:
            comment = Comment(comment_text.strip(), "System")
            comment.width = 200
            comment.height = 400
            sheet.cell(row_num, comment_target_col_idx).comment = comment

    for cell in sheet["G"]:
        cell.number_format = '0'

    for col_idx_hide in range(13, 29):
        sheet.column_dimensions[get_column_letter(col_idx_hide)].hidden = True

    sheet.cell(1, 29).value = "Menü"

    starts_with_list = ["1 4", "2 4", "3 4", "4 4", "5 4", "6 4"]
    source_count_col_idx = col_l_idx
    target_count_col_idx = 29

    for row_num in range(2, sheet.max_row + 1):
        cell_value_j = sheet.cell(row_num, source_count_col_idx).value
        count = 0
        if cell_value_j and isinstance(cell_value_j, str):
            lines = cell_value_j.splitlines()
            for line in lines:
                for start_pattern in starts_with_list:
                    if line.strip().startswith(start_pattern):
                        try:
                            count += int(start_pattern.split()[0])
                        except (ValueError, IndexError):
                            pass
        sheet.cell(row_num, target_count_col_idx).value = count if count > 0 else None

    current_date_str = datetime.utcnow().strftime("%Y%m%d")
    modified_file_name = f"{current_date_str}.xlsx"
    os.makedirs(save_directory, exist_ok=True)
    modified_file_path = os.path.join(save_directory, modified_file_name)
    workbook.save(modified_file_path)
    return modified_file_path


def build_clipboard_preview(sheet):
    clipboard_data_lines = []
    for row_num in range(2, sheet.max_row + 1):
        if sheet.cell(row_num, 6).value in [1, 2]:  # Spalte F
            row_data = []
            for col_idx in range(1, 12):  # bis einschließlich K
                cell_val = sheet.cell(row_num, col_idx).value
                if col_idx == 11:  # K
                    if cell_val and "LHK" in str(cell_val):
                        row_data.append("LHK")
                    else:
                        row_data.append("MS")
                else:
                    row_data.append(str(cell_val) if cell_val is not None else "")
            clipboard_data_lines.append("\t".join(row_data))
    return "\n".join(clipboard_data_lines)


@app.route("/telematik/process", methods=["POST"])
def telematik_process():
    try:
        excel = request.files.get("excel")
        if not excel:
            return jsonify({"status": "error", "message": "Excel-Datei fehlt (Feldname: excel)."}), 400

        # Upload speichern
        upload_name = excel.filename or "telematik.xlsx"
        upload_path = os.path.join(UPLOAD_FOLDER, upload_name)
        excel.save(upload_path)

        # Verarbeiten
        result_path = process_excel(upload_path, RESULTS_FOLDER)
        if not result_path or not os.path.exists(result_path):
            return jsonify({"status": "error", "message": "Ergebnisdatei wurde nicht erstellt."}), 500

        # Clipboard-Preview aus dem Ergebnis erzeugen
        wb = openpyxl.load_workbook(result_path)
        sheet = wb.active
        clipboard_preview = build_clipboard_preview(sheet)

        download_name = os.path.basename(result_path)
        return jsonify({
            "status": "ok",
            "message": "Telematik-Datei verarbeitet.",
            "download": f"/results/{download_name}",
            "clipboardPreview": clipboard_preview
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e), "trace": traceback.format_exc()}), 500


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        if email.lower() == ADMIN_EMAIL.lower() and password == ADMIN_PASSWORD:
            session["user_email"] = email
            session.permanent = True
            session["last_activity"] = int(datetime.utcnow().timestamp())
            return redirect(url_for("index"))
        return ("""<!doctype html><html><body>
                   <p>Zugangsdaten ungültig.</p>
                   <a href="/py/login">Zurück</a>
                   </body></html>""", 401)
    if session.get("user_email"):
        return redirect(url_for("index"))
    return """<!doctype html><html><body>
              <form method="post">
                <label>Email: <input name="email" /></label><br/>
                <label>Password: <input name="password" type="password"/></label><br/>
                <button type="submit">Login</button>
              </form>
              </body></html>"""


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def index():
    return """<!doctype html><html><body>
              <h3>Oflaz Mieten-Backend</h3>
              <p>Dieses Backend stellt Programmierschnittstellen bereit.</p>
              </body></html>"""


@app.route("/process", methods=["POST"])
def process():
    excel = request.files.get("excel")
    konto_file = request.files.get("konto")

    if not excel or not konto_file:
        return jsonify({"status": "error", "message": "Bitte Excel (Mieter) und Excel (Kontoauszug) hochladen."}), 400

    excel_path = os.path.join(UPLOAD_FOLDER, excel.filename)
    konto_path = os.path.join(UPLOAD_FOLDER, konto_file.filename)

    excel.save(excel_path)
    konto_file.save(konto_path)

    try:
        result_path = fuehre_mietabgleich_durch(excel_path, konto_path)
        if not result_path or not os.path.exists(result_path):
            return jsonify({"status": "error", "message": "Ergebnisdatei wurde nicht erstellt."}), 500

        download_name = os.path.basename(result_path)
        return jsonify({
            "status": "ok",
            "message": "Mietabgleich abgeschlossen",
            "download": f"/results/{download_name}"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e),
            "trace": traceback.format_exc()
        }), 500


@app.route("/results/<path:filename>")
def download_result(filename):
    file_path = os.path.join(RESULTS_FOLDER, filename)
    if not os.path.exists(file_path):
        return jsonify({"status": "error", "message": "Datei nicht gefunden"}), 404
    return send_from_directory(
        RESULTS_FOLDER,
        filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name=filename
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5005, debug=True, use_reloader=True)


