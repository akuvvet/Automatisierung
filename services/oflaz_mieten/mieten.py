import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.cell.cell import MergedCell
import os
import re
from pandas.api.types import is_datetime64_any_dtype, is_numeric_dtype
from datetime import datetime
from openpyxl.utils import get_column_letter

MONATS_ZUORDNUNG = {
    "Jan": ["Jan", "ZE-Jan"],
    "Feb": ["Feb", "ZE-Feb"],
    "Mrz": ["Mrz", "ZE-Mrz"],
    "Apr": ["Apr", "ZE-Apr"],
    "Mai": ["Mai", "ZE-Mai"],
    "Jun": ["Jun", "ZE-Jun"],
    "Jul": ["Jul", "ZE-Jul"],
    "Aug": ["Aug", "ZE-Aug"],
    "Sep": ["Sep", "ZE-Sep"],
    "Okt": ["Okt", "ZE-Okt"],
    "Nov": ["Nov", "ZE-Nov"],
    "Dez": ["Dez", "ZE-Dez"],
}

MONATS_NAMENS_MAPPING = {
    "Januar": "Jan", "Jan": "Jan",
    "Februar": "Feb", "Feb": "Feb",
    "März": "Mrz", "Marz": "Mrz", "Mrz": "Mrz",
    "April": "Apr", "Apr": "Apr",
    "Mai": "Mai",
    "Juni": "Jun", "Jun": "Jun",
    "Juli": "Jul", "Jul": "Jul",
    "August": "Aug", "Aug": "Aug",
    "September": "Sep", "Sep": "Sep", "Sept": "Sep",
    "Oktober": "Okt", "Okt": "Okt",
    "November": "Nov", "Nov": "Nov",
    "Dezember": "Dez", "Dez": "Dez",
}

BLATTNAME = "mieter"
MIETER_SPALTE = "A"

KONTO_DATUM = "Wertstellung"
KONTO_PAYEE = "Empfänger/Auftraggeber"
KONTO_VWZ = "Verwendungszweck"
KONTO_KATEGORIE = "Kategorie"  # optional
KONTO_OBJEKT = "Kontoname"
KONTO_BETRAG = "Betrag"


def fuehre_mietabgleich_durch(excel_pfad, konto_xlsx_pfad):
    workbook = load_workbook(excel_pfad)
    if BLATTNAME in workbook.sheetnames:
        try:
            worksheet = workbook[BLATTNAME]
        except Exception:
            return None
        sheet_arg = BLATTNAME
    else:
        first_sheet = workbook.sheetnames[0]
        worksheet = workbook[first_sheet]
        sheet_arg = first_sheet

    def _build_header_map(ws, max_scan_rows: int = 5) -> dict[str, str]:
        for r in range(1, min(ws.max_row, max_scan_rows) + 1):
            mapping: dict[str, str] = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                name = str(v).strip()
                if name and name not in mapping:
                    mapping[name] = get_column_letter(c)
            if mapping:
                return mapping
        return {}

    try:
        header_map = _build_header_map(worksheet)
    except Exception:
        header_map = {}

    df_mieter = pd.read_excel(excel_pfad, sheet_name=sheet_arg, dtype=str)
    df_mieter = df_mieter.fillna("")
    mieter_col_name = df_mieter.columns[0]
    mieter_b_col_name = df_mieter.columns[1] if len(df_mieter.columns) > 1 else None
    objekt_col_name = df_mieter.columns[2] if len(df_mieter.columns) > 2 else None

    mieter_row_map = {}
    try:
        for r in range(1, worksheet.max_row + 1):
            cell_val = worksheet[f"{MIETER_SPALTE}{r}"].value
            if cell_val is None:
                continue
            key = str(cell_val)
            key = re.sub(r"[^a-z0-9\s]", " ", key.lower())
            key = key.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
            key = re.sub(r"\s+", " ", key).strip()
            if key and key not in mieter_row_map:
                mieter_row_map[key] = r
    except Exception:
        pass

    try:
        df_konto = pd.read_excel(konto_xlsx_pfad, dtype=str)
    except Exception:
        return None

    df_konto = df_konto.fillna("")

    required_cols = [KONTO_DATUM, KONTO_PAYEE, KONTO_VWZ, KONTO_OBJEKT, KONTO_BETRAG]
    missing = [c for c in required_cols if c not in df_konto.columns]
    if missing:
        raise ValueError(f"Fehlende Spalten im Kontoauszug: {', '.join(missing)}. Bitte per Überschrift bereitstellen.")

    df_konto["__betrag_raw"] = df_konto[KONTO_BETRAG].astype(str)

    def _parse_amount_raw(raw: str):
        s = (str(raw) or "").strip().replace("€", "").replace("\xa0", "").replace(" ", "")
        if not s:
            return pd.NA
        if "," in s:
            s2 = s.replace(".", "").replace(",", ".")
            try:
                return float(s2)
            except Exception:
                pass
        try:
            return float(s)
        except Exception:
            m = re.search(r"(\d+)[,\.](\d{1,2})$", s)
            if m:
                try:
                    return float(m.group(1) + "." + m.group(2))
                except Exception:
                    pass
        return pd.NA

    df_konto[KONTO_BETRAG] = df_konto["__betrag_raw"].apply(_parse_amount_raw)

    s = df_konto[KONTO_DATUM]
    if is_datetime64_any_dtype(s):
        df_konto["__raw_date"] = s.dt.strftime("%d.%m.%Y")
    elif is_numeric_dtype(s):
        dt = pd.to_datetime(s, unit="d", origin="1899-12-30", errors="coerce")
        df_konto["__raw_date"] = dt.dt.strftime("%d.%m.%Y")
        df_konto[KONTO_DATUM] = dt
    else:
        s2 = (s.astype(str).str.strip().str.replace(r"\s+", "", regex=True).str.replace("/", ".", regex=False))
        df_konto["__raw_date"] = s2
        df_konto[KONTO_DATUM] = pd.to_datetime(s2, format="%d.%m.%Y", errors="coerce")

    def klassifiziere(text: str) -> str:
        t = (text or "").lower()
        patterns = [
            ("Miete", [r"\bmiet\w*\b", r"\bkm\b", r"\bkaltmiete\b", r"\bstellplatz\b", r"\bgarage\b"]),
            ("Nebenkosten", [r"\bnebenkosten\b", r"\bnk\b", r"\bbetriebskosten\b", r"\bbk\b", r"\bhausgeld\b", r"\bheizkosten\b"]),
            ("Nachzahlung", [r"\bnach\-?zahlung\b", r"\bnachz\b"]),
            ("Rate", [r"\brate(nzahlung)?\b"]),
            ("Honorar", [r"\bhonorar\b"]),
        ]
        for label, pats in patterns:
            for p in pats:
                if re.search(p, t, re.IGNORECASE):
                    return label
        return "Sonstiges"

    cat_series = df_konto[KONTO_KATEGORIE].astype(str) if KONTO_KATEGORIE in df_konto.columns else pd.Series([""] * len(df_konto), index=df_konto.index)
    df_konto["__text_summe"] = (df_konto[KONTO_VWZ].astype(str) + " " + cat_series + " " + df_konto[KONTO_OBJEKT].astype(str))
    df_konto["__klass"] = df_konto["__text_summe"].apply(klassifiziere)

    def finde_suchwort(text: str) -> str:
        t = (text or "").lower()
        patterns = [
            (None, [r"\bmiet\w*\b", r"\bkm\b", r"\bkaltmiete\b", r"\bstellplatz\b", r"\bgarage\b"]),
            (None, [r"\bnebenkosten\b", r"\bnk\b", r"\bbetriebskosten\b", r"\bbk\b"]),
            (None, [r"\bnach\-?zahlung\b", r"\bnachz\b"]),
            (None, [r"\brate(nzahlung)?\b"]),
            (None, [r"\bhonorar\b"]),
        ]
        for _, pats in patterns:
            for p in pats:
                m = re.search(p, t, re.IGNORECASE)
                if m:
                    return m.group(0)
        return ""

    df_konto["__hit"] = df_konto["__text_summe"].apply(finde_suchwort)

    month_key_map = {k.lower(): v for k, v in MONATS_NAMENS_MAPPING.items()}
    month_pattern = re.compile(r"\b(" + "|".join(re.escape(k) for k in MONATS_NAMENS_MAPPING.keys()) + r")\b", flags=re.IGNORECASE)

    def finde_monats_override(vwz_text: str) -> str | None:
        if not vwz_text:
            return None
        m = month_pattern.search(vwz_text)
        if not m:
            return None
        return month_key_map.get(m.group(1).lower())

    cat_for_override = df_konto[KONTO_KATEGORIE].astype(str) if KONTO_KATEGORIE in df_konto.columns else pd.Series([""] * len(df_konto), index=df_konto.index)
    df_konto["__month_override"] = ((df_konto[KONTO_VWZ].astype(str) + " " + cat_for_override).apply(finde_monats_override))

    sheet_such = "suchtreffer"
    if sheet_such in workbook.sheetnames:
        del workbook[sheet_such]
    ws_such = workbook.create_sheet(sheet_such)
    ws_such.append(["Datum", "Name", "Suchwort", "Betrag", "Zielmonat"])

    relevante_labels = {"Miete", "Nebenkosten", "Nachzahlung", "Rate", "Honorar"}
    df_such = df_konto[(df_konto["__klass"].isin(relevante_labels)) | (df_konto["__month_override"].notna())].copy()
    try:
        df_such = df_such.sort_values([KONTO_PAYEE, KONTO_DATUM, KONTO_BETRAG], kind="mergesort")
    except Exception:
        pass

    for _, r in df_such.iterrows():
        hit = r["__hit"] if r["__hit"] else r["__klass"]
        ziel_monat = r.get("__month_override", "")
        ws_such.append(["", r[KONTO_PAYEE], str(hit), r[KONTO_BETRAG], ziel_monat if ziel_monat else ""])
        row_idx = ws_such.max_row
        date_cell = ws_such.cell(row=row_idx, column=1)
        raw_val = r.get("__raw_date", "")
        dval = r[KONTO_DATUM]
        try:
            if pd.notna(dval):
                py_dt = dval.to_pydatetime() if hasattr(dval, "to_pydatetime") else dval
                date_cell.value = py_dt.date()
                date_cell.number_format = "DD.MM.YYYY"
            else:
                rv = (str(raw_val) or "").strip()
                m_iso = re.search(r"(\d{4})[-/\.](\d{2})[-/\.](\d{2})", rv)
                if m_iso:
                    parsed = datetime(int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3)))
                    date_cell.value = parsed.date()
                    date_cell.number_format = "DD.MM.YYYY"
                else:
                    rv2 = rv.replace("/", ".")
                    parsed = datetime.strptime(rv2, "%d.%m.%Y")
                    date_cell.value = parsed.date()
                    date_cell.number_format = "DD.MM.YYYY"
        except Exception:
            date_cell.value = None
            date_cell.number_format = "DD.MM.YYYY"
        try:
            ws_such.cell(row=row_idx, column=4).number_format = "#,##0.00"
        except Exception:
            pass

    months_order = list(MONATS_ZUORDNUNG.keys())

    def _get_writable_cell(ws, coord):
        c = ws[coord]
        if isinstance(c, MergedCell):
            for r in ws.merged_cells.ranges:
                if coord in r:
                    return ws.cell(row=r.min_row, column=r.min_col)
        return c

    def _write_date(cell, dt_val, raw_val) -> str:
        if pd.notna(dt_val):
            try:
                py_dt = dt_val.to_pydatetime() if hasattr(dt_val, "to_pydatetime") else dt_val
                cell.value = py_dt.date()
                cell.number_format = "DD.MM.YYYY"
                return py_dt.strftime("%d.%m.%Y")
            except Exception:
                pass
        rv = (str(raw_val) or "").strip()
        m_iso = re.search(r"(\d{4})[-/\.](\d{2})[-/\.](\d{2})", rv)
        try:
            if m_iso:
                parsed = datetime(int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3)))
            else:
                parsed = datetime.strptime(rv.replace("/", "."), "%d.%m.%Y")
            cell.value = parsed.date()
            cell.number_format = "DD.MM.YYYY"
            return parsed.strftime("%d.%m.%Y")
        except Exception:
            cell.value = None
            cell.number_format = "DD.MM.YYYY"
            return ""

    def _parse_amount_cell(val) -> float:
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).replace(".", "").replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    def _norm_ddmmyyyy(s: str) -> str:
        s = (s or "").strip()
        m = re.match(r"^(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?$", s)
        if not m:
            return s
        d = int(m.group(1)); mth = int(m.group(2)); yr = m.group(3)
        if yr is None:
            return f"{d:02d}.{mth:02d}"
        return f"{d:02d}.{mth:02d}.{int(yr):04d}"

    def _parse_pairs(txt: str):
        pairs = []
        seen = set()
        if not txt:
            return pairs
        for line in txt.splitlines():
            m = re.search(r"(\d{1,2}\.\d{1,2}(?:\.\d{2,4})?)\s*(?:\[(.*?)\])?\s*:\s*([+-]?\d+(?:[.,]\d+)?)", line)
            if not m:
                continue
            d = _norm_ddmmyyyy(m.group(1))
            try:
                amt = round(float(m.group(3).replace(".", "").replace(",", ".")), 2)
            except Exception:
                continue
            key = (d, amt)
            if key in seen:
                continue
            seen.add(key)
            pairs.append((d, amt, m.group(1), (m.group(2) or "").strip()))
        return pairs

    df_such["__norm_payee"] = df_such[KONTO_PAYEE].astype(str).apply(lambda val: re.sub(r"\s+", " ", re.sub(r"[^a-z0-9\s]", " ", str(val).lower()).replace("ä","ae").replace("ö","oe").replace("ü","ue").replace("ß","ss")).strip())

    for _, row in df_mieter.iterrows():
        m_name = row[mieter_col_name]
        if not m_name:
            continue
        key_norm = re.sub(r"\s+", " ", re.sub(r"[^a-z0-9\s]", " ", str(m_name).lower()).replace("ä","ae").replace("ö","oe").replace("ü","ue").replace("ß","ss")).strip()
        excel_row = mieter_row_map.get(key_norm)
        if not excel_row:
            continue

        owner_norm = key_norm
        tenant_norm = re.sub(r"\s+", " ", re.sub(r"[^a-z0-9\s]", " ", str(row[mieter_b_col_name] if mieter_b_col_name else "").lower()).replace("ä","ae").replace("ö","oe").replace("ü","ue").replace("ß","ss")).strip()
        GOV_KEYS = ("jobcenter", "agentur", "stadt wuppertal")
        if any(k in owner_norm for k in GOV_KEYS) and tenant_norm:
            treffer = df_such[df_such.get("__norm_hit", df_such["__norm_payee"]).str.contains(tenant_norm, na=False)]
        else:
            treffer = df_such[df_such["__norm_payee"] == owner_norm]
        if treffer.empty:
            continue

        for _, t in treffer.iterrows():
            dval = t[KONTO_DATUM]
            raw_date = t.get("__raw_date", "")
            kw = t["__hit"] if t["__hit"] else t["__klass"]
            betrag = t[KONTO_BETRAG]
            if pd.isna(betrag):
                continue

            month_idx = None
            override = t.get("__month_override", None)
            if isinstance(override, str) and override in months_order:
                month_idx = months_order.index(override)
            if pd.notna(dval):
                try:
                    month_idx = int(getattr(dval, "month"))
                except Exception:
                    month_idx = None
            if month_idx is None:
                rv = str(raw_date or "").strip()
                m_iso = re.search(r"(\d{4})[-/\.](\d{2})[-/\.](\d{2})", rv)
                if m_iso:
                    try:
                        month_idx = int(m_iso.group(2))
                    except Exception:
                        month_idx = None
                if month_idx is None:
                    try:
                        month_idx = datetime.strptime(rv, "%d.%m.%Y").month
                    except Exception:
                        month_idx = None
            if month_idx is None:
                continue
            month_idx = month_idx - 1
            if month_idx < 0 or month_idx > 11:
                continue
            ziel = list(MONATS_ZUORDNUNG.keys())[month_idx]

            betrag_hdr, datum_hdr = MONATS_ZUORDNUNG[ziel]
            betrag_letter = header_map.get(betrag_hdr)
            datum_letter = header_map.get(datum_hdr)
            if not betrag_letter or not datum_letter:
                continue
            betrag_cell = f"{betrag_letter}{excel_row}"
            datum_cell = f"{datum_letter}{excel_row}"

            date_cell = _get_writable_cell(worksheet, datum_cell)
            prev_cell_val = date_cell.value
            existing_amount = 0.0
            v = worksheet[betrag_cell].value
            if isinstance(v, (int, float)):
                existing_amount = float(v)
            else:
                try:
                    existing_amount = float(str(v).replace(".", "").replace(",", ".")) if v else 0.0
                except Exception:
                    existing_amount = 0.0

            existing_pairs = _parse_pairs(date_cell.comment.text if date_cell.comment else "")
            existing_keys = {(d, a) for (d, a, _, __) in existing_pairs}

            if pd.notna(dval):
                try:
                    new_date_str = dval.strftime("%d.%m.%Y")
                except Exception:
                    new_date_str = ""
            else:
                rv = str(raw_date or "").strip()
                m_iso = re.search(r"(\d{4})[-/\.](\d{2})[-/\.](\d{2})", rv)
                if m_iso:
                    try:
                        new_date_str = datetime(int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3))).strftime("%d.%m.%Y")
                    except Exception:
                        new_date_str = rv
                else:
                    try:
                        new_date_str = datetime.strptime(rv, "%d.%m.%Y").strftime("%d.%m.%Y")
                    except Exception:
                        new_date_str = rv
            def _norm_ddmmyyyy_local(s: str) -> str:
                s = (s or "").strip()
                m = re.match(r"^(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?$", s)
                if not m:
                    return s
                d = int(m.group(1)); mth = int(m.group(2)); yr = m.group(3)
                if yr is None:
                    return f"{d:02d}.{mth:02d}"
                return f"{d:02d}.{mth:02d}.{int(yr):04d}"
            new_key = (_norm_ddmmyyyy_local(new_date_str), round(float(betrag), 2))
            if new_key in existing_keys:
                continue
            if (existing_amount > 0.0) and not existing_pairs:
                prev_str = prev_cell_val.strftime("%d.%m.%Y") if hasattr(prev_cell_val, "strftime") else (str(prev_cell_val) if prev_cell_val else "")
                prev_key = (_norm_ddmmyyyy_local(prev_str), round(existing_amount, 2))
                if prev_key == new_key:
                    continue

            worksheet[betrag_cell] = existing_amount + float(betrag)
            try:
                worksheet[betrag_cell].number_format = "#,##0.00"
            except Exception:
                pass
            # Datum setzen (Format bleibt erhalten; bei leerem/ungültigem Datum wird None gesetzt)
            try:
                py_dt = t[KONTO_DATUM].to_pydatetime() if hasattr(t[KONTO_DATUM], "to_pydatetime") else t[KONTO_DATUM]
                if pd.notna(py_dt):
                    date_cell.value = py_dt.date()
                    date_cell.number_format = "DD.MM.YYYY"
                else:
                    date_cell.value = None
                    date_cell.number_format = "DD.MM.YYYY"
            except Exception:
                date_cell.value = None
                date_cell.number_format = "DD.MM.YYYY"

            has_previous_in_month = (existing_amount > 0.0) or (len(existing_pairs) > 0)
            if has_previous_in_month:
                lines = []
                if existing_pairs:
                    for (_, amt, disp, kword) in existing_pairs:
                        tag = f"{disp}"
                        if kword:
                            tag += f" [{kword}]"
                        lines.append(f"{tag}: {str(f'{amt:.2f}').replace('.', ',')} EUR")
                else:
                    prev_str = prev_cell_val.strftime("%d.%m.%Y") if hasattr(prev_cell_val, "strftime") else (str(prev_cell_val) if prev_cell_val else "")
                    if prev_str:
                        lines.append(f"{prev_str}: {str(f'{existing_amount:.2f}').replace('.', ',')} EUR")
                tag_new = new_date_str
                if kw:
                    tag_new += f" [{kw}]"
                lines.append(f"{tag_new}: {str(f'{float(betrag):.2f}').replace('.', ',')} EUR")
                date_cell.comment = Comment("\n".join(lines), "System")
            else:
                date_cell.comment = None

    result_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results", "mieten_abgleich.xlsx")
    os.makedirs(os.path.dirname(result_path), exist_ok=True)
    workbook.save(result_path)
    return result_path


